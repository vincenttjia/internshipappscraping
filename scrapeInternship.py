import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Cara Mendapatkan SSO Url
# Buka https://binus.ac.id/internship/dashboard/
# Klik Kanan Tombol Internship Apps
# Copy Link Address
SSOUrl="https://internship.apps.binus.ac.id/Login/Auth/SSOMicrosite?t=V1NC3NT-TJ14N4TT4N-SS0-T0K3N"

def main():
	session = requests.Session()
	page = session.get(SSOUrl)
	data = []

	soup = BeautifulSoup(page.content, 'html.parser')
	table = soup.find('table', attrs={'id':'JobList'})
	table_body = table.find('tbody')
	rows = table_body.find_all('tr')

	for row in rows:
		cols = row.find_all('td')
		cols = [ele.text.strip() for ele in cols]
		data.append([ele for ele in cols if ele])

	parse(data)

def parse(datas):
	i=0
	length = len(datas)
	workbook = Workbook()
	worksheet = workbook.worksheets[0]
	worksheet.title = "Sheet1"
	
	worksheet.cell(row=1, column=1).value = "Job Position"
	worksheet.cell(row=1, column=2).value = "Nama"
	worksheet.cell(row=1, column=3).value = "Contact Person"
	worksheet.cell(row=1, column=4).value = "Requirements"
	worksheet.cell(row=1, column=5).value = "Tempat Kerja"
	worksheet.cell(row=1, column=7).value = "Start"
	worksheet.cell(row=1, column=8).value = "Stop"
	worksheet.cell(row=1, column=9).value = "Length"
	worksheet.cell(row=1, column=11).value = "Type"
	worksheet.cell(row=1, column=12).value = "Quota"
	while(i<length):
		index=i+2
		data = datas[i]
		data = data[1]
		data = data.replace("\r\n","<vincent>")
		data1 = data.split("\n")
		worksheet.cell(row=index, column=1).value = data1[4].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=2).value = data1[0].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=3).value = data1[2].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=4).value = data1[11].strip("<vincent>").strip().replace("<vincent>","\n")
		worksheet.cell(row=index, column=5).value = data1[18].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=7).value = data1[28].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=8).value = data1[32].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=9).value = data1[22].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=11).value = data1[38].replace("<vincent>","").strip()
		worksheet.cell(row=index, column=12).value = data1[42].replace("<vincent>","").strip()
		i+=1

	workbook.save('output.xlsx')


if __name__ == '__main__':
	main()
